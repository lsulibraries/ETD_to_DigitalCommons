#! /usr/bin/env python3

import os
import csv
from collections import namedtuple
from datetime import datetime
import urllib.request

import openpyxl


# these items' files were chpt-by-chpt pdfs, joined into one pdf post upload.
# we must be careful to ingest to Digital Commons the joined files instead of the split files.

joined_postupload = ('etd-06182004-122626', 'etd-09012004-114224', 'etd-0327102-091522', 'etd-0707103-142120',
                     'etd-0710102-054039', 'etd-0409103-184148', 'etd-04152004-142117', 'etd-0830102-145811',
                     'etd-0903103-141852', )


def list_all_sheets(workbook):
    # non-essential sanity check function
    sheets = [sheet for sheet in available_wb.get_sheet_names()]
    print(sheets)


def show_all_info(urn):
    # non-essential script that prints all the info related to a urn.
    print("Main example:", main_sheet[urn], "\n")
    print("Filenames example:", filenames_sheet[urn], "\n")
    print("Keywords example:", keywords_sheet[urn], "\n")
    print("Advisors example:", advisors_sheet[urn], "\n")
    print("Catalog example:", catalog_sheet[urn], "\n")


def show_combinations_of_advisors(advisors_sheet):
    urn_advisortitles = dict()
    for urn, advisors_nt_list in advisors_sheet.items():
        for item in advisors_nt_list:
            if item.urn in urn_advisortitles:
                urn_advisortitles[item.urn].append(item.advisor_title)
            else:
                urn_advisortitles[item.urn] = [item.advisor_title, ]

    a_set = set()
    for urn, titles in urn_advisortitles.items():
        for title in titles:
            a_set.add(title)
    print(a_set)

    advisors_permutations = set()

    for urn, titles in urn_advisortitles.items():
        this_permutation = (titles.count('Committee Chair'),
                            titles.count('Committee Co-Chair'),
                            titles.count('Committee Member'),
                            titles.count("Dean's Representative"),
                            )
        advisors_permutations.add(this_permutation)
    for i in advisors_permutations:
        print(i)
    return advisors_permutations


def find_mismatching_files(filenames_sheet):
    sames = dict()
    for urn, filenames_namedtuple_list in filenames_sheet.items():
        for item in filenames_namedtuple_list:
            if item.urn in sames:
                if sames[item.urn] != item.availability:
                    print('there should be one {}'.format(item.urn))
            else:
                sames[item.urn] = item.availability
    return sames


def find_misnamed_extensions(filenames_sheet):
    misnamed_urn_filename = []
    for urn, filenames_namedtuple_list in filenames_sheet.items():
        for item in filenames_namedtuple_list:
            if item.filename[-4] != "." and item.filename[-4:] not in ("docx", "r.gz"):
                misnamed_urn_filename.append((urn, item.filename))
                print(urn, item.filename)
    return misnamed_urn_filename


def find_legacy_school_names(main_sheet):
    schools_etds = dict()
    for urn, itemnamedtuple in main_sheet.items():
        if itemnamedtuple.department in schools_etds:
            schools_etds[itemnamedtuple.department].append(urn)
        else:
            schools_etds[itemnamedtuple.department] = [urn, ]
    for school, urns in schools_etds.items():
        print(school)
    return schools_etds


def find_page_by_page_pdfs(filenames_sheet):
    split_files = dict()
    for urn, filenames_namedtuples_list in filenames_sheet.items():
        for item in filenames_namedtuples_list:
            if item.urn not in split_files:
                split_files[item.urn] = [item.filename, ]
            else:
                split_files[item.urn].append(item.filename)
    page_by_page_pdfs = []
    for urn, filelist in split_files.items():
        split = False
        for i in filelist:
            if "chap" in i.lower():
                split = True
        if len(filelist) > 1 and split == True:
            print(urn, '\n', filelist, '\n')
            page_by_page_pdfs.append((urn, filelist))
    return page_by_page_pdfs


def is_catalog_superset_of_database(catalog_sheet, main_sheet):
    outside_uris = []
    for uri in catalog_sheet:
        uri = os.path.split(uri)[1]
        if uri not in main_sheet:
            # print(uri)
            outside_uris.append(uri)
    print(len(outside_uris))
    return outside_uris


def parse_workbook(workbook_name):
    sourcepath = 'data/databasetables'
    filename = 'prod_etd_{}_database.xlsx'.format(workbook_name)
    fullpath = os.path.join(sourcepath, filename)
    return openpyxl.load_workbook(fullpath)


def parse_main_sheet():
    """ returns a dictionary in form of:
    {urn: NamedTuple
     urn: NamedTuple
    }
    NamedTuple is expected to have attributes: (urn first_name middle_name last_name suffix author_email
                                                publish_email degree department dtype title abstract availability
                                                availability_description copyright_statement ddate sdate adate
                                                cdate rdate pid url notice notice_response timestamp
                                                survey_completed)
                                            or: (urn first_name middle_name last_name suffix author_email
                                                publish_email degree department dtype title abstract availability
                                                availability_description copyright_statement ddate sdate adate
                                                cdate rdate pid url notices timestamp)
    """
    main_dict = dict()
    for wb in (available_wb, submitted_wb, withheld_wb):
        current_sheet = wb.get_sheet_by_name('etd_main table')
        for num, row in enumerate(current_sheet.iter_rows()):
            if num == 0:
                headers = (i.value for i in row)
                MainSheet = namedtuple('MainSheet', headers)
                continue
            values = (i.value for i in row)
            item = MainSheet(*values)
            main_dict[item.urn] = item
    return main_dict


def parse_filename_sheet():
    """ returns a dictionary in form of:
        urn: [NamedTuple, NamedTuple, ],
        urn: [NamedTuple, ]
    NamedTuple is expected to have attributes (path, size, available, description, page_count, timestamp)
    """
    filenames_sheet = dict()
    for wb in (available_wb, submitted_wb, withheld_wb):
        current_sheet = wb.get_sheet_by_name('filename_by_urn table')
        for num, row in enumerate(current_sheet.iter_rows()):
            if num == 0:
                headers = (i.value for i in row)
                Filenames = namedtuple('Filenames', headers)
                continue
            values = (i.value for i in row)
            item = Filenames(*values)

            if item.urn not in filenames_sheet:
                filenames_sheet[item.urn] = [item, ]
            else:
                row_timestamp = datetime.strptime(item.timestamp, "%Y-%m-%d %H:%M:%S")
                if item.filename in filenames_sheet[item.urn]:
                    previous_filename_entry = [i for i in filenames_sheet[item.urn] if i.filename == item.filename]
                    previous_timestamp = datetime.strptime(previous_filename_entry[0].timestamp, "%Y-%m-%d %H:%M:%S")
                    if row_timestamp > previous_timestamp:
                        print('oops')
                        previous_filename_entry[0] = item
                else:
                    filenames_sheet[item.urn].append(item)
    return filenames_sheet


def parse_keyword_sheet():
    """ returns a dictionary in form of:
    {urn: [NamedTuple,
           NamedTuple,
           ]}
    NamedTuple is expected to have attributes ('keyword', 'urn', 'timestamp')
    """
    keywords_sheet = dict()
    for wb in (available_wb, submitted_wb, withheld_wb):
        current_sheet = wb.get_sheet_by_name('keyword_by_urn table')
        for num, row in enumerate(current_sheet.iter_rows()):
            if num == 0:
                headers = (i.value for i in row)
                Keywords = namedtuple('Keywords', headers)
                continue
            values = (i.value for i in row)
            item = Keywords(*values)
            if item.urn not in keywords_sheet:
                keywords_sheet[item.urn] = [item, ]
            else:
                keywords_sheet[item.urn].append(item)
    return keywords_sheet


def parse_advisors_sheet():
    """ returns a dictionary in form of:
        {urn: [NamedTuple,
               NamedTuple,
               ]}
        NamedTuple is expected to have attributes ('urn', 'advisor_name', 'advisor_title',
                                                   'advisor_email', 'approval', 'timestamp')
   """
    advisors_sheet = dict()
    for wb in (available_wb, submitted_wb, withheld_wb):
        current_sheet = wb.get_sheet_by_name('advisor_by_urn table')
        for num, row in enumerate(current_sheet.iter_rows()):
            if num == 0:
                headers = (i.value for i in row)
                Advisors = namedtuple('Advisor', headers)
                continue
            values = (i.value for i in row)
            item = Advisors(*values)
            if item.urn not in advisors_sheet:
                advisors_sheet[item.urn] = [item, ]
            else:
                row_timestamp = datetime.strptime(item.timestamp, "%Y-%m-%d %H:%M:%S")
                if item.advisor_name in advisors_sheet[item.urn]:
                    previous_advisor_entry = [i for i in advisors_sheet[item.urn] if i.advisor_name == item.advisor_name]
                    previous_timestamp = datetime.strptime(previous_advisor_entry[0].timestamp, "%Y-%m-%d %H:%M:%S")
                    if row_timestamp > previous_timestamp:
                        previous_advisor_entry[0] = item
                else:
                    advisors_sheet[item.urn].append(item)
    return advisors_sheet


def parse_catalog_sheet():
    """ returns a dictionary in form of:
        {urn: [NamedTuple, NamedTuple, ]
         urn: [NamedTuple, ]}
    """
    catalog_sheet = dict()
    sourcepath = 'data/Catalogtables'
    sourcefile = 'CatalogETDSelectMetadata.csv'
    with open(os.path.join(sourcepath, sourcefile)) as csvfile:
        csvreader = csv.reader(csvfile, delimiter=',')
        for num, row in enumerate(csvreader):
            if num == 0:
                headers = (i for i in row)
                Catalog = namedtuple('Catalog', headers)
                continue
            values = (i for i in row)
            item = Catalog(*values)
            urn = [i for i in os.path.split(item.URL) if 'etd-' in i]
            urn = urn[0]
            if not urn:
                print('No urn for URL:', item.URL)
            if urn not in catalog_sheet:
                catalog_sheet[urn] = [item, ]
            else:
                catalog_sheet[urn].append(item)
    return catalog_sheet


def retrieve_binary(url):
    with urllib.request.urlopen(url) as response:
        return response.read()


def write_binary_to_file(binary, folder, filename):
    os.makedirs(folder, exist_ok=True)
    filepath = os.path.join(folder, filename)
    with open(filepath, 'bw') as f:
        f.write(binary)


def scrape_binaries(filenames_sheet):
    didnt_grab = []
    target_dir = './ETDbinaries/'
    count = 0
    for urn, filenames_namedtuples_list in filenames_sheet.items():
        local_dir = os.path.join(target_dir, urn)
        local_files = []
        if os.path.isdir(local_dir):
            local_files = os.listdir(local_dir)
        for item in filenames_namedtuples_list:
            if item.filename in local_files:
                pass
            else:
                url = 'http://etd.lsu.edu/{}/{}'.format("/".join(item.path.split('/')[3:]),
                                                                 item.filename)
                try:
                    binary = retrieve_binary(url)
                    write_binary_to_file(binary, local_dir, item.filename)
                except:
                    count += 1
                    pass
                    didnt_grab.append((urn, item.filename))
                    print(urn, item.filename)
    print(count)
    return didnt_grab


def csv_writer(data, path):
    with open(path, "w", newline='') as csv_file:
        writer = csv.writer(csv_file, delimiter='\t')
        for line in data:
            writer.writerow(line)


def concatinate_keywords(keywords_sheet, urn):
    return ', '.join(nt.keyword for nt in keywords_sheet[urn] if nt.keyword)


def organize_advisors(advisors_sheet, urn):
    Advisors = advisors_sheet[urn]
    advisors_rank = {'Committee Chair': 1, 'Committee Co-Chair': 2, 'Committee Member': 3, "Dean's Representative": 4}
    alpha_Advisors = sorted(Advisors, key=lambda x: x.advisor_name)
    sorted_advisors = sorted(alpha_Advisors, key=lambda x: advisors_rank[x.advisor_title])
    if len(sorted_advisors) > 7 and "Dean's Representative" in sorted_advisors[-1]:
        sorted_advisors = sorted_advisors[:6] + sorted_advisors[-1:]
    elif len(sorted_advisors) > 7:
        sorted_advisors = sorted_advisors[:7]
    return sorted_advisors


def 

available_wb = parse_workbook('available')
submitted_wb = parse_workbook('submitted')
withheld_wb = parse_workbook('withheld')

# merges the matching sheets from all 3 workbooks into one datastructure per sheet-type.
main_sheet = parse_main_sheet()
catalog_sheet = parse_catalog_sheet()
filenames_sheet = parse_filename_sheet()
keywords_sheet = parse_keyword_sheet()
advisors_sheet = parse_advisors_sheet()

# show_combinations_of_advisors(advisors_sheet)
# find_legacy_school_names(main_sheet)
